"""Xuất TKB ra Excel (Phase 5) — dùng openpyxl; dữ liệu lấy từ bảng `tkb` + service tkb.grid / grid_gv."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.models import GiaoVien, Lop
from app.services import tkb as tkb_svc

DAY_NHAN = {2: 'Thứ 2', 3: 'Thứ 3', 4: 'Thứ 4', 5: 'Thứ 5', 6: 'Thứ 6', 7: 'Thứ 7', 8: 'Chủ nhật'}

BOLD = Font(bold=True)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _sheets(session, sheets):
    """sheets: list of (title, lop_id_or_None, gv_id_or_None, note)."""
    wb = Workbook()
    wb.remove(wb.active)
    for title, lop_id, gv_id, note in sheets:
        if lop_id is not None:
            rows, tiet_labels, days = tkb_svc.grid(session, lop_id)
        else:
            rows, tiet_labels, days = tkb_svc.grid_gv(session, gv_id)
        ws = wb.create_sheet(title=title[:31])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(days))
        ws.cell(1, 1, note).font = BOLD
        # header
        for c, day in enumerate(days, start=2):
            cell = ws.cell(2, c, DAY_NHAN[day])
            cell.font = BOLD
            cell.alignment = CENTER
        ws.cell(2, 1, 'Tiết').font = BOLD
        # data
        for r, tl in enumerate(tiet_labels, start=3):
            ws.cell(r, 1, tl)
            for c, day in enumerate(days, start=2):
                v = rows[tl].get(DAY_NHAN[day], '') or ''
                cell = ws.cell(r, c, v)
                cell.alignment = CENTER
        ws.column_dimensions['A'].width = 14
        for c in range(2, 2 + len(days)):
            ws.column_dimensions[get_column_letter(c)].width = 18
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def xuat_lop(session, lop_id) -> BytesIO:
    lop = session.query(Lop).get(lop_id)
    return _sheets(session, [(f'TKB {lop.ten}', lop_id, None, f'Thời khóa biểu lớp {lop.ten}')])


def xuat_gv(session, gv_id) -> BytesIO:
    gv = session.query(GiaoVien).get(gv_id)
    return _sheets(session, [(f'TKB GV {gv.ten[:20]}', None, gv_id, f'Lịch giảng dạy: {gv.ten}')])


def xuat_toan_truong(session) -> BytesIO:
    sheets = [(l.ten, l.id, None, f'Thời khóa biểu lớp {l.ten}')
              for l in session.query(Lop).order_by(Lop.khoi_id, Lop.ten).all()]
    return _sheets(session, sheets)