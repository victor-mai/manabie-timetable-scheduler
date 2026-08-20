"""Import/Export Excel (openpyxl) + thao tác file SQLite."""

from io import BytesIO

from openpyxl import Workbook, load_workbook

# ánh xạ cột cho mỗi loại dữ liệu (tên cột tiếng Việt người dùng nhìn thấy)
TEMPLATE = {
    'mon':        {'ma': 'Mã môn', 'ten': 'Tên môn'},
    'gv':         {'ma': 'Mã GV', 'ten': 'Họ tên'},
    'khoi':       {'ten': 'Tên khối'},
    'lop':        {'ten': 'Lớp', 'khoi': 'Khối', 'si_so': 'Sĩ số', 'co_so': 'Cơ sở'},
}


def _rows(df):
    """pandas DataFrame -> list of dict (cột gốc)."""
    return df.to_dict('records')


def build_workbook(sheets: dict[str, list[dict]]) -> BytesIO:
    """sheets: {tên sheet: [dict dòng]} -> BytesIO xlsx."""
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        if not rows:
            ws.append([])
            continue
        cols = list(rows[0].keys())
        ws.append(cols)
        for r in rows:
            ws.append([r.get(c, '') for c in cols])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def read_excel(data: bytes):
    """Đọc file xlsx bất kỳ -> {sheet_name: [dict rows]}."""
    wb = load_workbook(BytesIO(data), data_only=True)
    out = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c) if c is not None else '' for c in rows[0]]
        out[ws.title] = [dict(zip(header, r)) for r in rows[1:] if any(r)]
    return out